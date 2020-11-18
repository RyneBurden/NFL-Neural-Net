#!/usr/bin/env python

# -------------------------------------------------- #
#
# Author: Ryne Burden
#
# Description:
#     - This script makes the data sets for training, validation during training, and in-season prediction data
#
# Tested on Python 3.8.1
#
# -------------------------------------------------- #

import openpyxl
import sys
import numpy as np
import teams
import os

# This function combines game data as is from competing teams for use in training Staley
def makeTrainingData(year, week):

    # Open the schedule for the given year to aid in data making
    wb = openpyxl.load_workbook('Data/Schedule/' + str(year) + '/' + str(year) + '.xlsx')

    # current sheet we are working with
    currentSheet = wb.active
    
    # Number of rows in the sheet
    numGames = currentSheet.max_row

    # This is the file object that the dataset will be written to
    outputFile = open('Data/Data Sets/' + str(year) + '/Training Data/' + str(year) + '_Week_' + str(week) + '.txt', 'w')
    print ('----- Now starting ' + str(year) + ' week ' + str(week) + ' -----\n')
    
    # Go through the whole sheet and find the games from the week given by the user
    for game in range(1, numGames + 1):
        # Variables to help parsing in the loop below
        gameWeek = 'A' + str(game)
        awayTeam = 'B' + str(game)
        homeTeam = 'C' + str(game)
        
        # Make data if the games being parsed is from the given week
        if (int(currentSheet[gameWeek].value) == int(week)):
            # Load all the season data for the given team(s) to extract what's needed
            # Replace spacesd with _ to help in loading data files (another _ needed when actually calling the txt file)
            awayFilename = (currentSheet[awayTeam].value).replace(" ", "_")
            homeFilename = (currentSheet[homeTeam].value).replace(" ", "_")
    
            # load the away team's whole season set into a numpy array for use later
            awayTeamMasterList = np.loadtxt('Data/Stats/' + str(year) + '/' + awayFilename + '_' + str(year) + '.txt')            
            # load the home team's whole season set into a numpy array for use later
            homeTeamMasterList = np.loadtxt('Data/Stats/' + str(year) + '/' + homeFilename + '_' + str(year) + '.txt')
            
            # Delete the last column to omit win / loss data for the away team
            # The home team win / loss will remain as the last value in our array as the expected value for our network
            awayTeamMasterList = np.delete(awayTeamMasterList, 8, axis=1)
            
            # Combine the desired weeks together
            # Take 1 from value of currentSheet[gameWeek].value to account for array index mapping
            currentGameData = np.concatenate((awayTeamMasterList[(int(currentSheet[gameWeek].value) - 1),:], homeTeamMasterList[(int(currentSheet[gameWeek].value) - 1),:]))
            
            # Write currentGameData to the outputFile 
            for entry in currentGameData:
                outputFile.write(str(entry) + ' ')
            outputFile.write('\n')
            
            # Print that the entry has been written successfully
            print(currentSheet[awayTeam].value + ' @ ' + currentSheet[homeTeam].value + ' ' + str(year) + ' Week ' + str(week) + ' written to file')

    print('\n----- ' + str(year) + ' week ' + str(week) + ' successfully writen -----\n')
    outputFile.close

# This function makes data based on the same schedule as the trainingData, but uses stat averages from the previous season to see how accurate the model is with unseen data that has significant figures
# I'm not sure how scientific this is but it seemed useful
# This function will be less useful in the future and probably not used after a normal season. 
# This is used to make data for the first 3 weeks in a season, using data from the previous season
def makeValidationData(year, week):

    # Open the schedule for the given year to aid in data making
    wb = openpyxl.load_workbook('Data/Schedule/' + str(year) + '/' + str(year) + '.xlsx')

    # current sheet we are working with
    currentSheet = wb.active
    
    # Number of rows in the sheet
    numGames = currentSheet.max_row

    # This is the file object that the dataset will be written to
    outputFile = open('Data/Data Sets/' + str(year) + '/Validation Data/' + str(int(year) - 1) + ' Averages/' + str(year) + '_Week_' + str(week) + 'v.txt', 'w')
        
    # Go through the whole sheet and find the games from the week given by the user
    for game in range(1, numGames + 1):
        # Variables to help parsing in the loop below
        gameWeek = 'A' + str(game)
        awayTeam = 'B' + str(game)
        homeTeam = 'C' + str(game)
        
        # Make data if the games being parsed is from the given week
        if (int(currentSheet[gameWeek].value) == int(week)):
            # Load all the season data for the given team(s) to extract what's needed
            # Replace spacesd with _ to help in loading data files (another _ needed when actually calling the txt file)
            awayFilename = (currentSheet[awayTeam].value).replace(" ", "_")
            homeFilename = (currentSheet[homeTeam].value).replace(" ", "_")

            # Load the AFC and NFC lists into NP arrays for use later
            AFCnp = np.asarray(teams.AFC)
            NFCnp = np.asarray(teams.NFC)

            # Set awayTeamIndex and homeTeamIndex to a dummy value out of the realistic range
            awayTeamIndex = -2
            homeTeamIndex = -2
            
            # Initialize home and away variables to False 
            awayAFC = False
            awayNFC = False
            homeAFC = False
            homeNFC = False
            
            # This loop goes through every team in AFCnp and NFCnp to determine the home and away team
            for x in range(4):
                for y in range(4):
                    # The conditionals below determine the division of both the home and away team and change the respective booleans
                    if currentSheet[awayTeam].value == AFCnp[x][y]:
                        # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                        # This is needed because the master list separates teams by row. This is the row number
                        awayTeamIndex = ((x*3) + (x+y))
                        awayAFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    elif currentSheet[awayTeam].value == NFCnp[x][y]:
                        awayTeamIndex = ((x*3) + (x+y))
                        awayNFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    if currentSheet[homeTeam].value == AFCnp[x][y]:
                        homeTeamIndex = ((x*3) + (x+y))
                        homeAFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    elif currentSheet[homeTeam].value == NFCnp[x][y]:
                        homeTeamIndex = ((x*3) + (x+y))
                        homeNFC = True

            # Load the previous season averages into numpy arrays
            avgMasterListAFC = np.loadtxt('Data/Data Sets/' + str(int(year) - 1) + '/Validation Data/' + str(int(year) - 1) + 'AFCavg.txt')
            avgMasterListNFC = np.loadtxt('Data/Data Sets/' + str(int(year) - 1) + '/Validation Data/' + str(int(year) - 1) + 'NFCavg.txt')

            # Combine team data for games based on the home and away booleans determined earlier using the indexes found in the nested loop above
            if awayAFC == True and homeAFC == True:
                currentGameData = np.concatenate((avgMasterListAFC[awayTeamIndex], avgMasterListAFC[homeTeamIndex]), axis = 0)
            elif awayAFC == True and homeNFC == True:
                currentGameData = np.concatenate((avgMasterListAFC[awayTeamIndex], avgMasterListNFC[homeTeamIndex]), axis = 0)
            elif awayNFC == True and homeNFC == True:
                currentGameData = np.concatenate((avgMasterListNFC[awayTeamIndex], avgMasterListNFC[homeTeamIndex]), axis = 0)
            elif awayNFC == True and homeAFC == True:
                currentGameData = np.concatenate((avgMasterListNFC[awayTeamIndex], avgMasterListAFC[homeTeamIndex]), axis = 0)

            # Write currentGameData to the output file
            for entry in currentGameData:
                outputFile.write(str(entry) + ' ')
            outputFile.write('\n')
 
            # Print statement to confirm which games has just been written. Helps confirm correct data
            print(currentSheet[awayTeam].value + ' @ ' + currentSheet[homeTeam].value + ' ' + str(year) + ' Week ' + str(week) + ' written to file')
            
    # Print statement that visually separates weeks when the script is run
    print('\n----- ' + str(year) + ' week ' + str(week) + ' successfully writen -----\n')

    # Close the output file for the currentWeek so a new one can be opened on the next function call
    outputFile.close

# This function will be used each week during the season to calculate season averages for each team and use those averages to make data sets for use in predictions
def makePredictionData(year, week):

    # Run averageMaker.py to update season averages before making data sets
    os.system('./average_maker.py ' + str(year))

    # Open the schedule for the given year to aid in data making
    wb = openpyxl.load_workbook('Data/Schedule/' + str(year) + '/' + str(year) + '.xlsx')

    # current sheet we are working with
    currentSheet = wb.active
    
    # Number of rows in the sheet
    numGames = currentSheet.max_row

    # This is the file object that the dataset will be written to
    outputFile = open('Data/Data Sets/' + str(year) + '/Validation Data/' + str(year) + '_Week_' + str(week) + 'v.txt', 'w')
    
    # Go through the whole sheet and find the games from the week given by the user
    for game in range(1, numGames + 1):
        # Variables to help parsing in the loop below
        gameWeek = 'A' + str(game)
        awayTeam = 'B' + str(game)
        homeTeam = 'C' + str(game)
        
        # Make data if the games being parsed is from the given week
        if (int(currentSheet[gameWeek].value) == int(week)):
            # Load all the season data for the given team(s) to extract what's needed
            # Replace spacesd with _ to help in loading data files (another _ needed when actually calling the txt file)
            awayFilename = (currentSheet[awayTeam].value).replace(" ", "_")
            homeFilename = (currentSheet[homeTeam].value).replace(" ", "_")

            # Load the AFC and NFC lists into NP arrays for use later
            AFCnp = np.asarray(teams.AFC)
            NFCnp = np.asarray(teams.NFC)

            # Set awayTeamIndex and homeTeamIndex to a dummy value out of the realistic range
            awayTeamIndex = -2
            homeTeamIndex = -2
            
            # Initialize home and away variables to False 
            awayAFC = False
            awayNFC = False
            homeAFC = False
            homeNFC = False
            
            # This loop goes through every team in AFCnp and NFCnp to determine the home and away team
            for x in range(4):
                for y in range(4):
                    # The conditionals below determine the division of both the home and away team and change the respective booleans
                    if currentSheet[awayTeam].value == AFCnp[x][y]:
                        # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                        # This is needed because the master list separates teams by row. This is the row number
                        awayTeamIndex = ((x*3) + (x+y))
                        awayAFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    elif currentSheet[awayTeam].value == NFCnp[x][y]:
                        awayTeamIndex = ((x*3) + (x+y))
                        awayNFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    if currentSheet[homeTeam].value == AFCnp[x][y]:
                        homeTeamIndex = ((x*3) + (x+y))
                        homeAFC = True
                    # This calculation gives us the index we need to search for in the AFC or NFC master average lists below
                    # This is needed because the master list separates teams by row. This is the row number
                    elif currentSheet[homeTeam].value == NFCnp[x][y]:
                        homeTeamIndex = ((x*3) + (x+y))
                        homeNFC = True

            # Load the previous season averages into numpy arrays
            avgMasterListAFC = np.loadtxt('Data/Data Sets/' + str(int(year)) + '/Validation Data/' + str(int(year)) + 'AFCavg.txt')
            avgMasterListNFC = np.loadtxt('Data/Data Sets/' + str(int(year)) + '/Validation Data/' + str(int(year)) + 'NFCavg.txt')

            # Combine team data for games based on the home and away booleans determined earlier using the indexes found in the nested loop above
            if awayAFC == True and homeAFC == True:
                currentGameData = np.concatenate((avgMasterListAFC[awayTeamIndex], avgMasterListAFC[homeTeamIndex]), axis = 0)
            elif awayAFC == True and homeNFC == True:
                currentGameData = np.concatenate((avgMasterListAFC[awayTeamIndex], avgMasterListNFC[homeTeamIndex]), axis = 0)
            elif awayNFC == True and homeNFC == True:
                currentGameData = np.concatenate((avgMasterListNFC[awayTeamIndex], avgMasterListNFC[homeTeamIndex]), axis = 0)
            elif awayNFC == True and homeAFC == True:
                currentGameData = np.concatenate((avgMasterListNFC[awayTeamIndex], avgMasterListAFC[homeTeamIndex]), axis = 0)

            # Write currentGameData to the output file
            for entry in currentGameData:
                outputFile.write(str(entry) + ' ')
            outputFile.write('\n')
 
            # Print statement to confirm which games has just been written. Helps confirm correct data
            print(currentSheet[awayTeam].value + ' @ ' + currentSheet[homeTeam].value + ' ' + str(year) + ' Week ' + str(week) + ' written to file')
            
    # Print statement that visually separates weeks when the script is run
    print('\n----- ' + str(year) + ' week ' + str(week) + ' successfully writen -----\n')

    # Close the output file for the currentWeek so a new one can be opened on the next function call
    outputFile.close



def main():

    # Get the year to make data for from the user
    currentYear = input("Year: ")

    # Get the data type fo by made from the user
    dataType = input("Make validation, training, or prediction data: ")
    
    # Give the past week to tune the loop below
    weekOrSeason = input("Make week data or season data sets: ")
    
    # Make weekly data
    if weekOrSeason == 'week':
        weekNum = input("What week are we making data for: ")
        if dataType == 't':
            makeTrainingData(currentYear, weekNum)
        elif dataType == 'p':
            print('\n')
            makePredictionData(currentYear, weekNum)
        elif dataType == 'v':
            makeValidationData(currentYear, weekNum)
    elif weekOrSeason == 'season':
    # This will be run if I want to make data sets for a whole season
        for currentWeek in range(1, 18):
            if dataType == 'v':
                makeValidationData(currentYear, currentWeek)
            elif dataType == 't':
                makeTrainingData(currentYear, currentWeek)
            elif dataType == 'p':
                makePredictionData(currentYear, currentWeek)

    # Simple print statement to end script run
    print ("\n****************************************************************\n\n")
    print("\t----- DONE -----\n\n")
    print ("\n****************************************************************")
    
main()
